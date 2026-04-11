"""
Formatador Excel para anotacao humana cega.

Le o CSV de anotacao cega (blind_annotation.csv) e gera uma copia
formatada em Excel (.xlsx) com:
  - Cabecalho colorido (azul escuro, fonte branca)
  - Colunas de dados em cinza claro
  - Colunas de anotacao humana em amarelo (destaque visual)
  - Dropdowns para classificacao (link_direto / texto_implicito / nenhum)
  - Freeze panes (cabecalho + video_id fixos)
  - Zebra striping para legibilidade
  - Aba README com instrucoes de preenchimento

Uso:
    python -m src.validation.excel_formatter
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# -- Constantes de estilo ---------------------------------------------------

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79",
                           fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

_DATA_FILL_ODD = PatternFill(start_color="F2F2F2", end_color="F2F2F2",
                             fill_type="solid")
_DATA_FILL_EVEN = PatternFill(start_color="FFFFFF", end_color="FFFFFF",
                              fill_type="solid")

_ANNOTATION_FILL_ODD = PatternFill(start_color="FFF2CC", end_color="FFF2CC",
                                   fill_type="solid")
_ANNOTATION_FILL_EVEN = PatternFill(start_color="FFFDE7", end_color="FFFDE7",
                                    fill_type="solid")

_NOTES_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF",
                          fill_type="solid")

_DATA_FONT = Font(name="Calibri", size=10, color="333333")
_ANNOTATION_FONT = Font(name="Calibri", size=10, bold=True, color="333333")

_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

_WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
_TOP_ALIGNMENT = Alignment(vertical="top")

# Mapeamento coluna -> largura
_COLUMN_WIDTHS = {
    'video_id': 14,
    'artist_name': 22,
    'title': 42,
    'youtube_url': 48,
    'youtube_channel_url': 48,
    'description': 60,
    'channel_bio': 60,
    'manual_call2go_video': 22,
    'manual_call2go_canal': 22,
    'manual_call2go_combinado': 24,
    'confianca': 14,
    'notas': 35,
}

# Colunas de anotacao humana (destaque amarelo)
_ANNOTATION_COLS = {
    'manual_call2go_video', 'manual_call2go_canal',
    'manual_call2go_combinado', 'confianca', 'notas'
}

# Colunas com wrap text (descricoes longas)
_WRAP_COLS = {'description', 'channel_bio', 'notas'}


def _create_readme_sheet(wb, census_mode=False):
    """Cria aba README com instrucoes de preenchimento."""
    ws = wb.create_sheet("README", 0)

    if census_mode:
        instructions = [
            ["INSTRUCOES PARA ANOTACAO HUMANA -- CENSO COMPLETO"],
            [""],
            ["Este arquivo contem TODOS os videos coletados do YouTube."],
            ["Voce deve classificar CADA video manualmente, sem consultar o detector."],
            [""],
            ["COLUNAS DE DADOS (cinza):"],
            ["  video_id           - Identificador unico do video no YouTube"],
            ["  artist_name        - Nome do artista"],
            ["  title              - Titulo do video"],
            ["  youtube_url        - Link direto para o video (abra no navegador se necessario)"],
            ["  youtube_channel_url- Link para o canal do artista no YouTube"],
            ["  description        - Descricao COMPLETA do video"],
            ["  channel_bio        - Bio COMPLETA do canal (descricao + links da aba Sobre)"],
            ["                       Links apos '---LINKS---' sao os links reais da aba Sobre"],
            ["                       Links com '[Canal Oficial]' vem do canal oficial (artistas com OAC)"],
            [""],
            ["COLUNAS DE ANOTACAO (amarelo) - PREENCHER:"],
            ["  manual_call2go_video    - A DESCRICAO DO VIDEO contem Call2Go?"],
            ["                            'SIM' = contem link Spotify ou CTA (ouca no Spotify)"],
            ["                            'NAO' = nao tem Call2Go na descricao"],
            ["  manual_call2go_canal    - A BIO DO CANAL contem Call2Go?"],
            ["                            mesma logica: SIM ou NAO"],
            ["  manual_call2go_combinado- QUALQUER fonte tem Call2Go?"],
            ["                            SIM se video OU canal = SIM"],
            ["  confianca               - Nivel de confianca: 'alta', 'media', ou 'baixa'"],
            ["  notas                   - Qualquer observacao relevante (opcional)"],
            [""],
            ["APOS CONCLUIR:"],
            ["  1. Salve como: data/validation/ground_truth.csv (File > Save As > CSV UTF-8)"],
            ["  2. Execute: python -m src.validation.cross_validator"],
            [""],
            ["IMPORTANTE: NAO consulte o detector automatizado antes de anotar."],
        ]
    else:
        instructions = [
            ["INSTRUCOES PARA ANOTACAO HUMANA CEGA"],
            [""],
            ["Este arquivo contem 91 videos da amostra adversarial estratificada."],
            ["Voce deve classificar CADA video manualmente, sem consultar o detector."],
            [""],
            ["COLUNAS DE DADOS (cinza):"],
            ["  video_id           - Identificador unico do video no YouTube"],
            ["  artist_name        - Nome do artista"],
            ["  title              - Titulo do video"],
            ["  youtube_url        - Link direto para o video (abra no navegador se necessario)"],
            ["  youtube_channel_url- Link para o canal do artista no YouTube"],
            ["  description        - Descricao COMPLETA do video"],
            ["  channel_bio        - Bio COMPLETA do canal (descricao + links da aba Sobre)"],
            ["                       Links apos '---LINKS---' sao os links reais da aba Sobre"],
            ["                       Links com '[Canal Oficial]' vem do canal oficial (artistas com OAC)"],
            [""],
            ["COLUNAS DE ANOTACAO (amarelo) - PREENCHER:"],
            ["  manual_call2go_video    - Classificacao da DESCRICAO DO VIDEO:"],
            ["                            'link_direto'     = contem link do Spotify"],
            ["                            'texto_implicito' = menciona Spotify como CTA"],
            ["                            'nenhum'          = sem Call2Go na descricao"],
            ["  manual_call2go_canal    - Classificacao da BIO DO CANAL:"],
            ["                            mesmas opcoes acima"],
            ["  manual_call2go_combinado- Classificacao FINAL:"],
            ["                            se video OU canal tem Call2Go -> combinado tambem tem"],
            ["                            prevalencia: link_direto > texto_implicito > nenhum"],
            ["  confianca               - Nivel de confianca: 'alta', 'media', ou 'baixa'"],
            ["  notas                   - Qualquer observacao relevante (opcional)"],
            [""],
            ["APOS CONCLUIR:"],
            ["  1. Salve como: data/validation/ground_truth.csv (File > Save As > CSV UTF-8)"],
            ["  2. Execute: python -m src.validation.cross_validator"],
            [""],
            ["IMPORTANTE: NAO consulte ground_truth_prefilled.csv (validacao circular)."],
        ]

    title_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    section_font = Font(name="Calibri", size=11, bold=True, color="333333")
    text_font = Font(name="Calibri", size=11, color="333333")
    warn_font = Font(name="Calibri", size=11, bold=True, color="CC0000")

    for i, row_data in enumerate(instructions, 1):
        cell = ws.cell(row=i, column=1, value=row_data[0] if row_data else "")
        if i == 1:
            cell.font = title_font
        elif row_data and row_data[0].startswith("COLUNAS") or \
                row_data and row_data[0].startswith("APOS"):
            cell.font = section_font
        elif row_data and row_data[0].startswith("IMPORTANTE"):
            cell.font = warn_font
        else:
            cell.font = text_font

    ws.column_dimensions['A'].width = 90
    return ws


def format_blind_annotation(
    input_csv="data/validation/blind_annotation.csv",
    output_xlsx="data/validation/blind_annotation.xlsx",
    census_mode=False
):
    """
    Le o CSV cego e gera versao Excel formatada para anotacao humana.

    Args:
        input_csv: Caminho do CSV de entrada (blind_annotation.csv).
        output_xlsx: Caminho do XLSX de saida formatado.
        census_mode: Se True, usa dropdowns SIM/NAO (censo completo).

    Returns:
        Caminho do arquivo gerado, ou None em caso de erro.
    """
    print("=" * 60)
    print("FORMATADOR EXCEL PARA ANOTACAO HUMANA")
    if census_mode:
        print("  (MODO CENSO -- dropdowns SIM/NAO)")
    print("=" * 60)

    if not os.path.exists(input_csv):
        print(f"[ERRO] CSV nao encontrado: {input_csv}")
        print("Execute primeiro: python -m src.validation.blind_annotator")
        return None

    # Carrega CSV
    df = pd.read_csv(input_csv, dtype=str, keep_default_na=False)
    print(f"  Videos carregados: {len(df)}")
    print(f"  Colunas: {list(df.columns)}")

    # Cria workbook
    wb = Workbook()

    # Aba README primeiro
    _create_readme_sheet(wb, census_mode=census_mode)

    # Remove sheet padrao e cria aba de dados
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    ws = wb.create_sheet("Anotacao", 1)

    columns = list(df.columns)

    # -- Cabecalho --
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _TOP_ALIGNMENT
        cell.border = _THIN_BORDER

    # -- Dados --
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        is_odd = (row_idx % 2 == 0)  # Zebra: linhas pares/impares alternadas
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[col_name])

            # Estilo baseado no tipo de coluna
            if col_name in _ANNOTATION_COLS:
                cell.fill = _ANNOTATION_FILL_ODD if is_odd else _ANNOTATION_FILL_EVEN
                cell.font = _ANNOTATION_FONT
            else:
                cell.fill = _DATA_FILL_ODD if is_odd else _DATA_FILL_EVEN
                cell.font = _DATA_FONT

            # Wrap text para descricoes longas
            if col_name in _WRAP_COLS:
                cell.alignment = _WRAP_ALIGNMENT
            else:
                cell.alignment = _TOP_ALIGNMENT

            cell.border = _THIN_BORDER

    # -- Larguras de coluna --
    for col_idx, col_name in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        width = _COLUMN_WIDTHS.get(col_name, 20)
        ws.column_dimensions[col_letter].width = width

    # -- Freeze panes: cabecalho fixo + video_id fixo ao rolar --
    ws.freeze_panes = "B2"

    # -- Auto-filtro no cabecalho --
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{len(df) + 1}"

    # -- Data validation (dropdowns) --
    last_row = len(df) + 1

    # Dropdown para colunas manual_call2go_*
    if census_mode:
        call2go_options = '"SIM,NAO"'
        call2go_error = "Escolha: SIM ou NAO"
        call2go_prompt = "Selecione SIM (tem Call2Go) ou NAO"
    else:
        call2go_options = '"link_direto,texto_implicito,nenhum"'
        call2go_error = "Escolha: link_direto, texto_implicito, ou nenhum"
        call2go_prompt = "Selecione a classificacao Call2Go"

    for col_name in ['manual_call2go_video', 'manual_call2go_canal',
                     'manual_call2go_combinado']:
        if col_name in columns:
            col_letter = get_column_letter(columns.index(col_name) + 1)
            dv = DataValidation(
                type="list",
                formula1=call2go_options,
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Valor invalido",
                error=call2go_error,
                showInputMessage=True,
                promptTitle=col_name,
                prompt=call2go_prompt
            )
            dv.add(f"{col_letter}2:{col_letter}{last_row}")
            ws.add_data_validation(dv)

    # Dropdown para confianca
    if 'confianca' in columns:
        col_letter = get_column_letter(columns.index('confianca') + 1)
        dv_conf = DataValidation(
            type="list",
            formula1='"alta,media,baixa"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Valor invalido",
            error="Escolha: alta, media, ou baixa",
            showInputMessage=True,
            promptTitle="confianca",
            prompt="Selecione o nivel de confianca"
        )
        dv_conf.add(f"{col_letter}2:{col_letter}{last_row}")
        ws.add_data_validation(dv_conf)

    # -- Altura das linhas (ajustar para descricoes longas) --
    for row_idx in range(2, last_row + 1):
        ws.row_dimensions[row_idx].height = 60

    # Salva
    os.makedirs(os.path.dirname(output_xlsx), exist_ok=True)
    wb.save(output_xlsx)

    print(f"\n  Excel formatado gerado: {output_xlsx}")
    print(f"  Abas: {wb.sheetnames}")
    print(f"  Linhas de dados: {len(df)}")
    print(f"  Dropdowns: manual_call2go_video, manual_call2go_canal, "
          f"manual_call2go_combinado, confianca")
    print(f"  Freeze panes: B2 (cabecalho + video_id fixos)")

    return output_xlsx


if __name__ == "__main__":
    import sys
    if '--census' in sys.argv:
        format_blind_annotation(
            input_csv="data/validation/blind_annotation_census.csv",
            output_xlsx="data/validation/blind_annotation_census.xlsx",
            census_mode=True
        )
    else:
        format_blind_annotation()
